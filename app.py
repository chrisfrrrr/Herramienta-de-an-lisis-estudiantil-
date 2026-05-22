import os
import re
import tempfile
from io import BytesIO
from datetime import datetime
from dataclasses import dataclass
from typing import List, Optional, Tuple

import fitz  # PyMuPDF
import pandas as pd
import streamlit as st
import plotly.express as px
from PIL import Image

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo

# =========================================================
# CONFIGURACIÓN GENERAL
# =========================================================
APP_TITLE = "Herramienta de análisis estudiantil - AVE"
CREDITS = "Desarrollado por Ing. Christian Pocol Asesor AVE"
SUBCREDITS = "Universidad del Valle de Guatemala UVG"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ASSETS_DIR = os.path.join(BASE_DIR, "assets")
LOGO_AVE = os.path.join(ASSETS_DIR, "logo_ave.png")
LOGO_UVG = os.path.join(ASSETS_DIR, "logo_uvg.png")

MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}

# =========================================================
# MODELO DE DATOS
# =========================================================
@dataclass
class Registro:
    user_id: str = ""
    nombre: str = ""
    login: str = ""
    sis: str = ""
    seccion: str = ""
    rol: str = ""
    ultima_actividad_texto: str = ""
    ultima_actividad_dt: Optional[datetime] = None
    actividad_total_texto: str = ""
    actividad_total_min: float = 0.0
    actividad_total_horas: float = 0.0
    horas_desde_ultima: Optional[float] = None
    dias_desde_ultima: Optional[float] = None
    riesgo_abandono: str = ""
    alerta_72h: str = ""
    avance_actividad: int = 0
    horas_esperadas: float = 0.0
    cumplimiento_semanal_pct: float = 0.0
    estado_cumplimiento: str = ""
    observacion: str = ""

# =========================================================
# FUNCIONES DE EXTRACCIÓN Y LIMPIEZA
# =========================================================
def limpiar_linea(s: str) -> str:
    s = s.replace("\u200b", "").replace("\ufeff", "").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def normalizar_pendiente(s: str) -> str:
    if re.fullmatch(r"p+e+n+d+i+e+n+t+e+", s.lower()):
        return "pendiente"
    return s


def extraer_texto_pdf(pdf_path: str) -> str:
    doc = fitz.open(pdf_path)
    partes = []
    for page in doc:
        partes.append(page.get_text("text"))
    doc.close()
    return "\n".join(partes)


def detectar_anio_reporte(texto: str, fecha_actual: datetime) -> int:
    # Ejemplo del pie de página del reporte: 21/5/26, 12:43
    m = re.search(r"\b(\d{1,2})/(\d{1,2})/(\d{2,4}),\s*\d{1,2}:\d{2}\b", texto)
    if m:
        y = int(m.group(3))
        return 2000 + y if y < 100 else y
    return fecha_actual.year


def segmentar_usuarios(texto: str) -> List[List[str]]:
    lineas = [limpiar_linea(x) for x in texto.splitlines()]
    lineas = [x for x in lineas if x]
    segmentos = []
    actual = []
    headers = {"Nombre", "Identificador de inicio", "de sesión", "Identificación", "del SIS", "Sección", "Rol", "Última", "actividad", "Actividad", "total"}

    for ln in lineas:
        if re.fullmatch(r"\(users/\d+\)", ln):
            if actual:
                segmentos.append(actual)
            actual = [ln]
        elif actual:
            if ln.startswith("https://uvg.instructure.com/courses"):
                continue
            if re.match(r"^\d{1,2}/\d{1,2}/\d{2,4},", ln):
                continue
            if ln in headers:
                continue
            actual.append(ln)
    if actual:
        segmentos.append(actual)
    return segmentos


def unir_nombre(nombre_lineas: List[str]) -> str:
    return limpiar_linea(" ".join(nombre_lineas))


def parse_login_sis(lineas: List[str], idx: int) -> Tuple[str, str, int]:
    if idx >= len(lineas):
        return "", "", idx
    ln = normalizar_pendiente(lineas[idx])

    m = re.match(r"(.+?@\S+)\s+(\d{3,})$", ln)
    if m:
        return m.group(1), m.group(2), idx + 1

    if ln.lower() == "pendiente":
        sis = ""
        if idx + 1 < len(lineas) and re.fullmatch(r"\d{3,}", lineas[idx + 1]):
            sis = lineas[idx + 1]
            idx += 1
        return "pendiente", sis, idx + 1

    if idx + 1 < len(lineas) and re.fullmatch(r"\d{3,}", lineas[idx + 1]):
        return ln, lineas[idx + 1], idx + 2

    return ln, "", idx + 1


def parse_actividad_total(texto: str) -> float:
    texto = (texto or "").strip()
    if not texto:
        return 0.0
    partes = texto.split(":")
    try:
        nums = [int(p) for p in partes]
    except Exception:
        return 0.0

    if len(nums) == 3:
        h, m, s = nums
        return h * 60 + m + s / 60
    if len(nums) == 2:
        # Canvas muestra duraciones menores a 1 hora como MM:SS
        m, s = nums
        return m + s / 60
    return 0.0


def parse_fecha_ultima(texto: str, anio: int) -> Optional[datetime]:
    if not texto:
        return None
    t = texto.lower()
    m = re.search(r"(\d{1,2})\s+de\s+([a-záéíóúñ]+)\s+en\s+(\d{1,2}):(\d{2})", t, re.IGNORECASE)
    if not m:
        return None

    dia = int(m.group(1))
    mes_txt = m.group(2).replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
    mes = MESES.get(mes_txt)
    if not mes:
        return None

    hora = int(m.group(3))
    minuto = int(m.group(4))
    try:
        return datetime(anio, mes, dia, hora, minuto)
    except ValueError:
        return None


def parse_segmento(seg: List[str], anio: int) -> Optional[Registro]:
    if not seg or not re.fullmatch(r"\(users/\d+\)", seg[0]):
        return None

    reg = Registro(user_id=re.search(r"\d+", seg[0]).group(0))

    url_idx = None
    for i, ln in enumerate(seg[1:], start=1):
        if ln.startswith("(https://"):
            url_idx = i
            break
    if url_idx is None:
        return None

    reg.nombre = unir_nombre(seg[1:url_idx])
    idx = url_idx + 1
    reg.login, reg.sis, idx = parse_login_sis(seg, idx)

    rol_idx = None
    for j in range(idx, len(seg)):
        if seg[j] in {"Estudiante", "Profesor", "Teacher", "Student"}:
            rol_idx = j
            break
    if rol_idx is None:
        return None

    reg.seccion = limpiar_linea(" ".join(seg[idx:rol_idx]))
    reg.rol = seg[rol_idx]
    resto = seg[rol_idx + 1:]

    tiempos = [x for x in resto if re.fullmatch(r"\d{1,2}:\d{2}(?::\d{2})?", x)]
    if len(tiempos) >= 2:
        reg.actividad_total_texto = tiempos[-1]
        pos_last_total = len(resto) - 1 - resto[::-1].index(tiempos[-1])
        reg.ultima_actividad_texto = limpiar_linea(" ".join(resto[:pos_last_total]))
    elif len(tiempos) == 1:
        joined = " ".join(resto).lower()
        if " de " in f" {joined} " or any(mes in joined for mes in MESES):
            reg.ultima_actividad_texto = limpiar_linea(" ".join(resto))
            reg.actividad_total_texto = ""
        else:
            reg.actividad_total_texto = tiempos[0]
            reg.ultima_actividad_texto = ""
    else:
        reg.ultima_actividad_texto = limpiar_linea(" ".join(resto))
        reg.actividad_total_texto = ""

    reg.actividad_total_min = round(parse_actividad_total(reg.actividad_total_texto), 2)
    reg.actividad_total_horas = round(reg.actividad_total_min / 60, 4)
    reg.ultima_actividad_dt = parse_fecha_ultima(reg.ultima_actividad_texto, anio)
    return reg

# =========================================================
# INDICADORES
# =========================================================
def calcular_indicadores(reg: Registro, semana: int, fecha_actual: datetime) -> Registro:
    reg.horas_esperadas = semana * 10.0

    if reg.ultima_actividad_dt:
        delta_h = (fecha_actual - reg.ultima_actividad_dt).total_seconds() / 3600
        reg.horas_desde_ultima = round(max(delta_h, 0), 2)
        reg.dias_desde_ultima = round(reg.horas_desde_ultima / 24, 2)

        if reg.horas_desde_ultima <= 24:
            reg.riesgo_abandono = "Bajo"
        elif reg.horas_desde_ultima <= 72:
            reg.riesgo_abandono = "Medio"
        else:
            reg.riesgo_abandono = "Alto"
        reg.alerta_72h = "Sí" if reg.horas_desde_ultima >= 72 else "No"
    else:
        reg.horas_desde_ultima = None
        reg.dias_desde_ultima = None
        reg.riesgo_abandono = "Alto"
        reg.alerta_72h = "Sin actividad registrada"

    if reg.actividad_total_min <= 10:
        reg.avance_actividad = 0
    elif reg.actividad_total_min <= 25:
        reg.avance_actividad = 25
    elif reg.actividad_total_min <= 120:
        reg.avance_actividad = 50
    else:
        reg.avance_actividad = 100

    reg.cumplimiento_semanal_pct = round(min((reg.actividad_total_horas / reg.horas_esperadas) * 100, 100), 2) if reg.horas_esperadas else 0
    reg.estado_cumplimiento = "Cumple" if reg.actividad_total_horas >= reg.horas_esperadas else "No cumple"

    obs = []
    if not reg.ultima_actividad_dt:
        obs.append("No registra última actividad")
    if reg.alerta_72h in {"Sí", "Sin actividad registrada"}:
        obs.append("Requiere seguimiento por inactividad")
    if reg.estado_cumplimiento == "No cumple":
        obs.append("No alcanza horas mínimas esperadas")
    reg.observacion = "; ".join(obs) if obs else "Sin alerta crítica"
    return reg


def procesar_pdf(pdf_path: str, semana: int, fecha_actual: Optional[datetime] = None) -> List[Registro]:
    fecha_actual = fecha_actual or datetime.now()
    texto = extraer_texto_pdf(pdf_path)
    anio = detectar_anio_reporte(texto, fecha_actual)
    segmentos = segmentar_usuarios(texto)

    registros = []
    for seg in segmentos:
        r = parse_segmento(seg, anio)
        if r and r.rol.lower() == "estudiante":
            registros.append(calcular_indicadores(r, semana, fecha_actual))
    return registros


def registros_a_dataframe(registros: List[Registro], semana: int) -> pd.DataFrame:
    rows = []
    for r in registros:
        rows.append({
            "ID Usuario": r.user_id,
            "Nombre": r.nombre,
            "Login": r.login,
            "SIS": r.sis,
            "Sección": r.seccion,
            "Rol": r.rol,
            "Última actividad": r.ultima_actividad_texto,
            "Fecha última actividad": r.ultima_actividad_dt.strftime("%Y-%m-%d %H:%M") if r.ultima_actividad_dt else "",
            "Actividad total": r.actividad_total_texto,
            "Actividad total (min)": r.actividad_total_min,
            "Actividad total (horas)": r.actividad_total_horas,
            "Horas desde última": r.horas_desde_ultima if r.horas_desde_ultima is not None else "",
            "Días desde última": r.dias_desde_ultima if r.dias_desde_ultima is not None else "",
            "Riesgo abandono": r.riesgo_abandono,
            "Alerta 72h": r.alerta_72h,
            "Avance actividad (%)": r.avance_actividad,
            "Semana análisis": semana,
            "Horas esperadas": r.horas_esperadas,
            "Cumplimiento semanal (%)": r.cumplimiento_semanal_pct,
            "Estado cumplimiento": r.estado_cumplimiento,
            "Observación": r.observacion,
        })
    return pd.DataFrame(rows)

# =========================================================
# EXCEL
# =========================================================
def aplicar_estilo_encabezado(ws, header_fill, header_font):
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def pintar_celda_por_riesgo(cell, valor):
    if valor == "Bajo":
        cell.fill = PatternFill("solid", fgColor="C6EFCE")
    elif valor == "Medio":
        cell.fill = PatternFill("solid", fgColor="FFEB9C")
    elif valor == "Alto":
        cell.fill = PatternFill("solid", fgColor="FFC7CE")


def ancho_columna_por_header(header: str) -> int:
    anchos = {
        "Archivo origen": 34,
        "ID Usuario": 12,
        "Nombre": 34,
        "Login": 26,
        "SIS": 12,
        "Sección": 32,
        "Rol": 12,
        "Última actividad": 24,
        "Fecha última actividad": 20,
        "Actividad total": 16,
        "Actividad total (min)": 18,
        "Actividad total (horas)": 20,
        "Horas desde última": 18,
        "Días desde última": 16,
        "Riesgo abandono": 16,
        "Alerta 72h": 22,
        "Avance actividad (%)": 18,
        "Semana análisis": 16,
        "Horas esperadas": 16,
        "Cumplimiento semanal (%)": 22,
        "Estado cumplimiento": 20,
        "Observación": 45,
    }
    return anchos.get(header, 18)


def aplicar_anchos(ws, headers):
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = ancho_columna_por_header(header)


def agregar_hoja_dataframe(wb, sheet_name: str, df_sheet: pd.DataFrame, header_fill, header_font, thin):
    ws = wb.create_sheet(sheet_name)
    headers = list(df_sheet.columns)
    ws.append(headers)
    for row in df_sheet.itertuples(index=False):
        ws.append(list(row))
    aplicar_estilo_encabezado(ws, header_fill, header_font)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    aplicar_anchos(ws, headers)
    idx_riesgo = headers.index("Riesgo abandono") + 1 if "Riesgo abandono" in headers else None
    idx_alerta = headers.index("Alerta 72h") + 1 if "Alerta 72h" in headers else None
    idx_cumple = headers.index("Estado cumplimiento") + 1 if "Estado cumplimiento" in headers else None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if idx_riesgo:
            pintar_celda_por_riesgo(row[idx_riesgo - 1], row[idx_riesgo - 1].value)
        if idx_alerta and row[idx_alerta - 1].value in {"Sí", "Sin actividad registrada"}:
            row[idx_alerta - 1].fill = PatternFill("solid", fgColor="FFC7CE")
        if idx_cumple and row[idx_cumple - 1].value == "No cumple":
            row[idx_cumple - 1].fill = PatternFill("solid", fgColor="FFC7CE")
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = Border(bottom=thin)
    return ws


def generar_resumen_por_seccion(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "Sección" not in df.columns:
        return pd.DataFrame()
    resumen = df.groupby("Sección", dropna=False).agg(
        Total_estudiantes=("Nombre", "count"),
        Con_actividad=("Fecha última actividad", lambda x: (x != "").sum()),
        Sin_actividad=("Fecha última actividad", lambda x: (x == "").sum()),
        Riesgo_bajo=("Riesgo abandono", lambda x: (x == "Bajo").sum()),
        Riesgo_medio=("Riesgo abandono", lambda x: (x == "Medio").sum()),
        Riesgo_alto=("Riesgo abandono", lambda x: (x == "Alto").sum()),
        Alertas_72h=("Alerta 72h", lambda x: x.isin(["Sí", "Sin actividad registrada"]).sum()),
        Cumplen=("Estado cumplimiento", lambda x: (x == "Cumple").sum()),
        No_cumplen=("Estado cumplimiento", lambda x: (x == "No cumple").sum()),
        Promedio_horas=("Actividad total (horas)", "mean"),
    ).reset_index()
    resumen["Promedio_horas"] = resumen["Promedio_horas"].round(2)
    resumen["% Riesgo alto"] = (resumen["Riesgo_alto"] / resumen["Total_estudiantes"] * 100).round(2)
    resumen["% Cumplimiento"] = (resumen["Cumplen"] / resumen["Total_estudiantes"] * 100).round(2)
    return resumen


def exportar_excel_bytes(df: pd.DataFrame, semana: int, fecha_actual: datetime, pdf_nombre: str) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte procesado"

    headers = list(df.columns)
    ws.append(headers)
    for row in df.itertuples(index=False):
        ws.append(list(row))

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    aplicar_estilo_encabezado(ws, header_fill, header_font)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    aplicar_anchos(ws, headers)

    idx_riesgo = headers.index("Riesgo abandono") + 1 if "Riesgo abandono" in headers else None
    idx_alerta = headers.index("Alerta 72h") + 1 if "Alerta 72h" in headers else None
    idx_cumple = headers.index("Estado cumplimiento") + 1 if "Estado cumplimiento" in headers else None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if idx_riesgo:
            pintar_celda_por_riesgo(row[idx_riesgo - 1], row[idx_riesgo - 1].value)
        if idx_alerta and row[idx_alerta - 1].value in {"Sí", "Sin actividad registrada"}:
            row[idx_alerta - 1].fill = PatternFill("solid", fgColor="FFC7CE")
        if idx_cumple and row[idx_cumple - 1].value == "No cumple":
            row[idx_cumple - 1].fill = PatternFill("solid", fgColor="FFC7CE")
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = Border(bottom=thin)

    if ws.max_row > 1 and ws.max_column > 1:
        last_col = get_column_letter(ws.max_column)
        tab = Table(displayName="TablaReporte", ref=f"A1:{last_col}{ws.max_row}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws.add_table(tab)

    # Hojas filtradas
    for sheet_name, filtro in [
        ("Alertas 72 horas", df["Alerta 72h"].isin(["Sí", "Sin actividad registrada"])),
        ("Riesgo alto", df["Riesgo abandono"].eq("Alto")),
    ]:
        agregar_hoja_dataframe(wb, sheet_name, df[filtro], header_fill, header_font, thin)

    # Resumen por sección
    resumen_seccion = generar_resumen_por_seccion(df)
    if not resumen_seccion.empty:
        wsr = wb.create_sheet("Resumen por sección")
        wsr.append(list(resumen_seccion.columns))
        for row in resumen_seccion.itertuples(index=False):
            wsr.append(list(row))
        aplicar_estilo_encabezado(wsr, header_fill, header_font)
        wsr.freeze_panes = "A2"
        wsr.auto_filter.ref = wsr.dimensions
        for col in range(1, wsr.max_column + 1):
            wsr.column_dimensions[get_column_letter(col)].width = 22
        wsr.column_dimensions["A"].width = 38

    # Estadísticas
    wse = wb.create_sheet("Estadísticas")
    total = len(df)
    activos = int((df["Fecha última actividad"] != "").sum()) if total else 0
    sin_act = total - activos
    bajo = int((df["Riesgo abandono"] == "Bajo").sum()) if total else 0
    medio = int((df["Riesgo abandono"] == "Medio").sum()) if total else 0
    alto = int((df["Riesgo abandono"] == "Alto").sum()) if total else 0
    alerta = int(df["Alerta 72h"].isin(["Sí", "Sin actividad registrada"]).sum()) if total else 0
    cumplen = int((df["Estado cumplimiento"] == "Cumple").sum()) if total else 0
    no_cumplen = total - cumplen
    promedio_horas = round(float(df["Actividad total (horas)"].mean()), 2) if total else 0
    total_secciones = int(df["Sección"].nunique()) if total and "Sección" in df.columns else 0

    wse["A1"] = "Resumen general del análisis"
    wse["A1"].font = Font(size=16, bold=True, color="1F4E78")
    resumen = [
        ("Archivos PDF analizados", pdf_nombre),
        ("Fecha/hora de análisis", fecha_actual.strftime("%Y-%m-%d %H:%M")),
        ("Semana seleccionada", semana),
        ("Horas esperadas acumuladas", semana * 10),
        ("Total de secciones detectadas", total_secciones),
        ("Total de estudiantes", total),
        ("Con última actividad registrada", activos),
        ("Sin actividad registrada", sin_act),
        ("Riesgo bajo", bajo),
        ("Riesgo medio", medio),
        ("Riesgo alto", alto),
        ("Alerta 72 horas o sin registro", alerta),
        ("Cumplen horas esperadas", cumplen),
        ("No cumplen horas esperadas", no_cumplen),
        ("Promedio actividad total (horas)", promedio_horas),
    ]
    for idx, (k, v) in enumerate(resumen, start=3):
        wse.cell(idx, 1).value = k
        wse.cell(idx, 2).value = v
        wse.cell(idx, 1).font = Font(bold=True)
    wse.column_dimensions["A"].width = 38
    wse.column_dimensions["B"].width = 50

    wse["D2"] = "Distribución por riesgo"
    wse["D2"].font = Font(bold=True)
    riesgo_data = [("Bajo", bajo), ("Medio", medio), ("Alto", alto)]
    for i, (cat, val) in enumerate(riesgo_data, start=3):
        wse.cell(i, 4).value = cat
        wse.cell(i, 5).value = val

    pie = PieChart()
    labels = Reference(wse, min_col=4, min_row=3, max_row=5)
    data = Reference(wse, min_col=5, min_row=2, max_row=5)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Riesgo de abandono"
    wse.add_chart(pie, "G2")

    wse["D8"] = "Cumplimiento semanal"
    wse["D8"].font = Font(bold=True)
    comp_data = [("Cumple", cumplen), ("No cumple", no_cumplen)]
    for i, (cat, val) in enumerate(comp_data, start=9):
        wse.cell(i, 4).value = cat
        wse.cell(i, 5).value = val
    bar = BarChart()
    labels2 = Reference(wse, min_col=4, min_row=9, max_row=10)
    data2 = Reference(wse, min_col=5, min_row=8, max_row=10)
    bar.add_data(data2, titles_from_data=True)
    bar.set_categories(labels2)
    bar.title = "Cumplimiento de horas esperadas"
    bar.y_axis.title = "Estudiantes"
    wse.add_chart(bar, "G18")

    # Criterios
    wsc = wb.create_sheet("Criterios")
    criterios = [
        ["Indicador", "Criterio aplicado"],
        ["Carga de archivos", "Permite cargar uno o varios PDF correspondientes a distintas secciones del mismo curso"],
        ["Consolidación", "Todos los PDF se unen en un análisis general y se conserva la columna Archivo origen"],
        ["Duplicados", "Si hay duplicados, se conserva el primer registro según SIS; si SIS está vacío, se usa ID Usuario"],
        ["Riesgo bajo", "0 a 24 horas desde la última actividad"],
        ["Riesgo medio", "Más de 24 y hasta 72 horas desde la última actividad"],
        ["Riesgo alto", "Más de 72 horas o sin actividad registrada"],
        ["Alerta 72h", "Sí cuando han pasado 72 horas o más desde la última conexión"],
        ["Horas esperadas", "Semana seleccionada x 10 horas acumuladas"],
        ["Avance 0%", "0 a 10 minutos de actividad total"],
        ["Avance 25%", "Más de 10 y hasta 25 minutos"],
        ["Avance 50%", "Más de 25 minutos y hasta 2 horas"],
        ["Avance 100%", "Más de 2 horas"],
    ]
    for row in criterios:
        wsc.append(row)
    aplicar_estilo_encabezado(wsc, header_fill, header_font)
    wsc.column_dimensions["A"].width = 24
    wsc.column_dimensions["B"].width = 88

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

# =========================================================
# INTERFAZ STREAMLIT
# =========================================================
st.set_page_config(page_title=APP_TITLE, page_icon="📊", layout="wide")

st.markdown(
    """
    <style>
    .main {background-color: #F6F8FB;}
    .block-container {padding-top: 1.5rem; padding-bottom: 2rem;}
    .ave-card {background: white; border-radius: 18px; padding: 22px; box-shadow: 0 4px 18px rgba(0,0,0,0.06); margin-bottom: 16px;}
    .ave-title {font-size: 2.0rem; font-weight: 800; color: #10216F; margin-bottom: 0.2rem;}
    .ave-credit {font-size: 1rem; color: #444; margin-bottom: 0.1rem;}
    .ave-subcredit {font-size: 1rem; color: #287A34; font-weight: 700;}
    .small-note {color: #666; font-size: 0.9rem;}
    </style>
    """,
    unsafe_allow_html=True,
)

# Header con logos
col_logo, col_title, col_uvg = st.columns([1.2, 3.6, 0.9])

with col_logo:
    st.markdown("<div style='height:55px;'></div>", unsafe_allow_html=True)
    if os.path.exists(LOGO_AVE):
        st.image(LOGO_AVE, width=190)

with col_title:
    st.markdown(
        f"""
        <div style='text-align:center; padding-top:10px;'>
            <div class='ave-title'>{APP_TITLE}</div>
            <div class='ave-credit'>{CREDITS}</div>
            <div class='ave-subcredit'>{SUBCREDITS}</div>
        </div>
        """,
        unsafe_allow_html=True
    )

with col_uvg:
    st.markdown("<div style='height:55px;'></div>", unsafe_allow_html=True)
    if os.path.exists(LOGO_UVG):
        st.image(LOGO_UVG, width=105)


st.markdown("---")

with st.sidebar:
    st.header("Configuración")
    semana = st.selectbox("Semana de análisis", [1, 2, 3, 4, 5], index=0)
    st.caption("Cada semana equivale a 10 horas acumuladas mínimas.")
    fecha_actual = datetime.now()
    st.info(f"Fecha/hora detectada por el sistema:\n\n{fecha_actual.strftime('%Y-%m-%d %H:%M')}")

    with st.expander("Criterios aplicados"):
        st.write("**Riesgo bajo:** 0 a 24 horas sin actividad.")
        st.write("**Riesgo medio:** más de 24 y hasta 72 horas.")
        st.write("**Riesgo alto:** más de 72 horas o sin registro.")
        st.write("**Avance 0%:** 0 a 10 minutos.")
        st.write("**Avance 25%:** más de 10 a 25 minutos.")
        st.write("**Avance 50%:** más de 25 minutos a 2 horas.")
        st.write("**Avance 100%:** más de 2 horas.")

st.markdown("<div class='ave-card'>", unsafe_allow_html=True)
st.subheader("Carga de reportes PDF")
st.write(
    "Sube uno o varios reportes PDF del mismo curso exportados desde Canvas/UVG. "
    "La aplicación consolidará todas las secciones y procesará únicamente los registros con rol **Estudiante**."
)
uploaded_files = st.file_uploader(
    "Selecciona uno o varios archivos PDF",
    type=["pdf"],
    accept_multiple_files=True,
)

if uploaded_files:
    st.success(f"Archivos cargados: {len(uploaded_files)}")
    with st.expander("Ver archivos cargados"):
        for file in uploaded_files:
            st.write(f"• {file.name}")

procesar = st.button("Procesar reportes", type="primary", use_container_width=True)
st.markdown("</div>", unsafe_allow_html=True)

if procesar:
    if not uploaded_files:
        st.error("Debe cargar uno o varios archivos PDF antes de procesar.")
    else:
        with st.spinner("Procesando PDFs, consolidando secciones y calculando indicadores..."):
            try:
                dataframes = []
                errores = []

                for uploaded_file in uploaded_files:
                    tmp_path = None
                    try:
                        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                            tmp.write(uploaded_file.getbuffer())
                            tmp_path = tmp.name

                        registros = procesar_pdf(tmp_path, int(semana), fecha_actual)

                        if registros:
                            df_temp = registros_a_dataframe(registros, int(semana))
                            df_temp.insert(0, "Archivo origen", uploaded_file.name)
                            dataframes.append(df_temp)
                        else:
                            errores.append(f"No se encontraron estudiantes válidos en: {uploaded_file.name}")
                    except Exception as e:
                        errores.append(f"Error procesando {uploaded_file.name}: {e}")
                    finally:
                        if tmp_path and os.path.exists(tmp_path):
                            os.remove(tmp_path)

                if not dataframes:
                    st.warning("No se encontraron estudiantes en los PDF cargados. Verifique que los archivos correspondan al formato del reporte del curso.")
                    for error in errores:
                        st.warning(error)
                else:
                    df = pd.concat(dataframes, ignore_index=True)

                    # Eliminar duplicados si un estudiante aparece en más de una sección o archivo.
                    # Prioridad: SIS; si SIS está vacío, se usa ID Usuario.
                    df["_clave_duplicado"] = df["SIS"].astype(str).str.strip()
                    df.loc[df["_clave_duplicado"].eq("") | df["_clave_duplicado"].str.lower().eq("nan"), "_clave_duplicado"] = df["ID Usuario"].astype(str)
                    duplicados = int(df.duplicated(subset=["_clave_duplicado"]).sum())
                    df = df.drop_duplicates(subset=["_clave_duplicado"], keep="first").drop(columns=["_clave_duplicado"])

                    nombres_archivos = "; ".join([file.name for file in uploaded_files])
                    excel_bytes = exportar_excel_bytes(df, int(semana), fecha_actual, nombres_archivos)

                    st.session_state["df_reporte"] = df
                    st.session_state["excel_bytes"] = excel_bytes.getvalue()
                    st.session_state["excel_name"] = f"analisis_general_curso_{fecha_actual.strftime('%Y%m%d_%H%M')}.xlsx"
                    st.session_state["semana"] = int(semana)
                    st.session_state["archivos_procesados"] = len(uploaded_files)
                    st.session_state["duplicados_eliminados"] = duplicados
                    st.session_state["errores_proceso"] = errores

                    st.success(
                        f"Análisis general generado correctamente. "
                        f"Archivos procesados: {len(uploaded_files)} · Estudiantes consolidados: {len(df)}"
                    )
                    if duplicados > 0:
                        st.info(f"Se eliminaron {duplicados} registro(s) duplicado(s) detectados por SIS o ID Usuario.")
                    for error in errores:
                        st.warning(error)
            except Exception as e:
                st.error(f"Ocurrió un error al procesar los PDF: {e}")

if "df_reporte" in st.session_state:
    df = st.session_state["df_reporte"]

    total = len(df)
    activos = int((df["Fecha última actividad"] != "").sum())
    sin_act = total - activos
    alerta = int(df["Alerta 72h"].isin(["Sí", "Sin actividad registrada"]).sum())
    riesgo_alto = int((df["Riesgo abandono"] == "Alto").sum())
    cumplen = int((df["Estado cumplimiento"] == "Cumple").sum())
    promedio_horas = round(float(df["Actividad total (horas)"].mean()), 2) if total else 0
    secciones_detectadas = int(df["Sección"].nunique()) if total else 0
    archivos_procesados = st.session_state.get("archivos_procesados", 1)

    st.subheader("Resumen general consolidado")
    k1, k2, k3, k4, k5, k6, k7 = st.columns(7)
    k1.metric("PDF procesados", archivos_procesados)
    k2.metric("Secciones", secciones_detectadas)
    k3.metric("Total estudiantes", total)
    k4.metric("Con actividad", activos)
    k5.metric("Sin actividad", sin_act)
    k6.metric("Riesgo alto", riesgo_alto)
    k7.metric("Alertas 72h", alerta)

    k8, k9, k10 = st.columns(3)
    k8.metric("Cumplen horas", cumplen)
    k9.metric("No cumplen", total - cumplen)
    k10.metric("Promedio horas", promedio_horas)

    if st.session_state.get("duplicados_eliminados", 0) > 0:
        st.info(f"Duplicados eliminados en la consolidación: {st.session_state['duplicados_eliminados']}")

    st.download_button(
        label="Descargar Excel consolidado con indicadores",
        data=st.session_state["excel_bytes"],
        file_name=st.session_state["excel_name"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    tab1, tab2, tab3, tab4, tab5 = st.tabs(["Reporte consolidado", "Resumen por sección", "Alertas", "Gráficas", "Estadísticas"])

    with tab1:
        st.dataframe(df, use_container_width=True, height=520)

    with tab2:
        resumen_seccion = generar_resumen_por_seccion(df)
        st.write("Resumen consolidado por sección detectada en los PDF cargados.")
        st.dataframe(resumen_seccion, use_container_width=True, height=420, hide_index=True)

    with tab3:
        df_alertas = df[df["Alerta 72h"].isin(["Sí", "Sin actividad registrada"]) | (df["Riesgo abandono"] == "Alto")].copy()
        st.write(f"Estudiantes con alerta o riesgo alto: **{len(df_alertas)}**")
        st.dataframe(df_alertas, use_container_width=True, height=500)

    with tab4:
        c1, c2 = st.columns(2)
        with c1:
            riesgo_counts = df["Riesgo abandono"].value_counts().reset_index()
            riesgo_counts.columns = ["Riesgo", "Cantidad"]
            fig = px.pie(riesgo_counts, names="Riesgo", values="Cantidad", title="Distribución general por riesgo de abandono")
            st.plotly_chart(fig, use_container_width=True)
        with c2:
            comp_counts = df["Estado cumplimiento"].value_counts().reset_index()
            comp_counts.columns = ["Estado", "Cantidad"]
            fig = px.bar(comp_counts, x="Estado", y="Cantidad", title="Cumplimiento general de horas esperadas", text="Cantidad")
            st.plotly_chart(fig, use_container_width=True)

        c3, c4 = st.columns(2)
        with c3:
            avance_counts = df["Avance actividad (%)"].value_counts().sort_index().reset_index()
            avance_counts.columns = ["Avance (%)", "Cantidad"]
            fig = px.bar(avance_counts, x="Avance (%)", y="Cantidad", title="Distribución general por avance de actividad", text="Cantidad")
            st.plotly_chart(fig, use_container_width=True)
        with c4:
            resumen_seccion = generar_resumen_por_seccion(df)
            if not resumen_seccion.empty:
                fig = px.bar(
                    resumen_seccion,
                    x="Sección",
                    y="Riesgo_alto",
                    title="Riesgo alto por sección",
                    text="Riesgo_alto",
                )
                st.plotly_chart(fig, use_container_width=True)

        top_inactividad = df.copy()
        top_inactividad["Horas desde última num"] = pd.to_numeric(top_inactividad["Horas desde última"], errors="coerce")
        top_inactividad = top_inactividad.sort_values("Horas desde última num", ascending=False).head(15)
        fig = px.bar(
            top_inactividad,
            x="Horas desde última num",
            y="Nombre",
            color="Sección" if "Sección" in top_inactividad.columns else None,
            orientation="h",
            title="Top 15 estudiantes con más horas sin actividad",
            hover_data=["Archivo origen", "SIS", "Sección"],
        )
        st.plotly_chart(fig, use_container_width=True)

    with tab5:
        stats = pd.DataFrame({
            "Indicador": [
                "Semana seleccionada",
                "Horas esperadas acumuladas",
                "PDF procesados",
                "Secciones detectadas",
                "Total de estudiantes",
                "Con última actividad registrada",
                "Sin actividad registrada",
                "Riesgo bajo",
                "Riesgo medio",
                "Riesgo alto",
                "Alerta 72 horas o sin registro",
                "Cumplen horas esperadas",
                "No cumplen horas esperadas",
                "Promedio actividad total (horas)",
                "Duplicados eliminados",
            ],
            "Valor": [
                st.session_state["semana"],
                st.session_state["semana"] * 10,
                archivos_procesados,
                secciones_detectadas,
                total,
                activos,
                sin_act,
                int((df["Riesgo abandono"] == "Bajo").sum()),
                int((df["Riesgo abandono"] == "Medio").sum()),
                riesgo_alto,
                alerta,
                cumplen,
                total - cumplen,
                promedio_horas,
                st.session_state.get("duplicados_eliminados", 0),
            ],
        })
        st.dataframe(stats, use_container_width=True, hide_index=True)

else:
    st.info("Carga uno o varios PDF, selecciona la semana de análisis y presiona **Procesar reportes** para generar los indicadores consolidados.")

st.markdown("---")
st.caption(f"{CREDITS} · {SUBCREDITS}")
